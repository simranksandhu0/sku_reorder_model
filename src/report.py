"""
report.py
Generates the final Excel reorder report with traffic-light status formatting.
Run this as the main entry point for the SKU reorder pipeline.

Usage:
    python src/report.py --input data/inventory.csv --output outputs/reorder_report.xlsx
"""

import argparse
import sys
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# Allow imports from project root when running src/report.py directly
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from src.velocity import load_and_validate, compute_velocity
from src.flags import load_thresholds, compute_days_coverage, apply_flags


STATUS_COLOURS = {
    "REORDER NOW": "FFCCCC",   # Red
    "SLOW MOVER":  "FFF2CC",   # Amber
    "WATCH":       "FFEB9C",   # Yellow
    "OK":          "CCFFCC",   # Green
}

OUTPUT_COLS = [
    "sku_id", "category", "stock_on_hand",
    "recent_velocity_daily", "prior_velocity_daily",
    "velocity_change_pct", "days_coverage",
    "status", "recommended_action",
]


def build_report_df(input_path: str) -> pd.DataFrame:
    """Run the full pipeline and return the report DataFrame."""
    df = load_and_validate(input_path)
    df = compute_velocity(df)
    df = compute_days_coverage(df)
    thresholds = load_thresholds()
    df = apply_flags(df, thresholds)

    # Sort: most urgent first
    status_order = {"REORDER NOW": 0, "SLOW MOVER": 1, "WATCH": 2, "OK": 3}
    df["_sort"] = df["status"].map(status_order)
    df = df.sort_values("_sort").drop(columns=["_sort"]).reset_index(drop=True)

    return df[OUTPUT_COLS]


def write_excel(df: pd.DataFrame, output_path: str) -> None:
    """Write the report DataFrame to a formatted Excel file."""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reorder Report"

    header_fill = PatternFill("solid", fgColor="2F5496")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Write header
    headers = list(df.columns)
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = border

    # Write data rows with traffic-light fills
    for row_data in dataframe_to_rows(df, index=False, header=False):
        ws.append(row_data)
        row_idx = ws.max_row
        status = ws.cell(row=row_idx, column=headers.index("status") + 1).value
        fill_colour = STATUS_COLOURS.get(status, "FFFFFF")
        fill = PatternFill("solid", fgColor=fill_colour)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)

    # Auto-size columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"Report saved to {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate SKU reorder report.")
    parser.add_argument("--input",  default="data/inventory.csv",
                        help="Path to inventory CSV")
    parser.add_argument("--output", default="outputs/reorder_report.xlsx",
                        help="Path for Excel output")
    args = parser.parse_args()

    df = build_report_df(args.input)

    print("\nStatus Summary:")
    print(df["status"].value_counts().to_string())

    write_excel(df, args.output)


if __name__ == "__main__":
    main()
